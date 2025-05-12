import tkinter as tk
from tkinter import ttk, filedialog, simpledialog
import pandas as pd
from pyvis.network import Network
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import random as random
import numpy as np
from copy import deepcopy
import math
from collections import defaultdict
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import Counter
import webbrowser
import os
import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTreeWidget, QTreeWidgetItem,
    QLineEdit, QPushButton, QVBoxLayout, QWidget, QLabel, QComboBox,
    QFileDialog, QHBoxLayout, QHeaderView, QMessageBox, QDialog
)
from PyQt5.QtCore import Qt, QUrl, QObject, pyqtSignal, QThread, QRunnable, QThreadPool, pyqtSlot, QTimer
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtGui import QColor, QBrush, QPixmap, QIcon
from PyQt5.QtWebChannel import QWebChannel
import tempfile
import json
import threading



# --------------CLUSTERING---------------
# Global variable to store parsed people
people = []
names = []

cluster_id = 1


def assign_clusters():
    global people, cluster_id

    visited = set()
    clusters = {}  # Dictionary to store persons assigned to each cluster

    # Inner DFS function to recursively assign clusterId
    def dfs(person):
        person['clusterId'] = cluster_id
        visited.add(person['name'])

        # Add the person to the current cluster's list
        clusters.setdefault(cluster_id, []).append(person)

        for friend_name in person['friends']:
            if not friend_name:  # Skip empty strings
                continue

            friend = next((p for p in people if p['name'] == friend_name), None)
            if friend and friend['name'] not in visited:
                dfs(friend)

    # Traverse each person in the list and start DFS if not visited
    for person in people:
        if person['name'] not in visited:
            clusters[cluster_id] = []  # Initialize a new list for the current cluster
            dfs(person)  # Initial DFS call

            # Perform additional checks for any indirect connections
            added_to_cluster = True
            while added_to_cluster:
                added_to_cluster = False
                for p in people:
                    if p['name'] not in visited:
                        if any(friend_name in visited for friend_name in p['friends'] if friend_name):
                            dfs(p)
                            added_to_cluster = True

            cluster_id += 1  # Increment clusterId for the next cluster

    return clusters


def get_first_and_second_names(name):
    name = name.split()
    return " ".join(name[:2])

def get_cluster_colors():
    # Identify all unique cluster IDs
    unique_cluster_ids = {person['clusterId'] for person in people}
    
    # Generate a unique color for each cluster ID
    cluster_colors = {}
    used_colors = set()  # To track already used colors

    for cluster_id in unique_cluster_ids:
        while True:
            # Generate a random color
            color_hex = ''.join(random.choices('0123456789ABCDEF', k=6))
            if color_hex not in used_colors:  # Ensure the color is unique
                break
        cluster_colors[cluster_id] = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        used_colors.add(color_hex)  # Mark the color as used

    return cluster_colors

def extract_names(data):
    names = set()
    for person in data:
        names.add(person["name"])
    return list(names)

def get_matches(name, names):
    matches = []
    name_lower = name.lower()
    match_size = len(name)

    # Generate matches until a match is found
    while match_size > 0:
        substrings = [
            name_lower[i:i + match_size] for i in range(len(name_lower) - match_size + 1)
        ]

        for candidate in names:
            candidate_lower = candidate.lower()
            for substring in substrings:
                if substring in candidate_lower:
                    matches.append(candidate)
            

        if len(matches) > 0:
                break
        match_size -= 1

    matches.append("None")

    return matches

def update_person(person, friend1_matches, friend2_matches):
    person['friends'][0] = "" if friend1_matches and friend1_matches[0] == "None" else friend1_matches[0] if friend1_matches else ""
    person['friends'][1] = "" if friend2_matches and friend2_matches[0] == "None" else friend2_matches[0] if friend2_matches else ""



    
# --------------FORMATION---------------

# This will store the number of teams required to be formed
number_of_teams = 0

initial_temp = 1000000.0
final_temp = 0.0
max_iterations = 10000
cooling_type = "linear"

# This will store the best distribution of teams so far in 2d array
best_teams_so_far = []

best_fitness_so_far = float('inf')

# This will store the current distribution of teams in 2d array
teams = []

def split_large_clusters(members):
    # Step 1: Group members by their clusters
    cluster_groups = defaultdict(list)
    for member in members:
        cluster_groups[member['clusterId']].append(member)

    # Step 2: Split clusters based on size
    new_members = []
    for cluster_id, group in cluster_groups.items():
        group_size = len(group)

        if group_size > 12:
            # Split into 3 clusters
            split_1 = group_size // 3
            split_2 = split_1 * 2

            cluster_1 = group[:split_1]
            cluster_2 = group[split_1:split_2]
            cluster_3 = group[split_2:]

            # Assign new cluster IDs
            for member in cluster_1:
                member['clusterId'] = f"{cluster_id}a"
            for member in cluster_2:
                member['clusterId'] = f"{cluster_id}b"
            for member in cluster_3:
                member['clusterId'] = f"{cluster_id}c"

            # Add modified members to the new list
            new_members.extend(cluster_1)
            new_members.extend(cluster_2)
            new_members.extend(cluster_3)

        elif group_size > 6:
            # Split into 2 clusters
            split_point = (group_size + 1) // 2  # Round up for the first half

            cluster_1 = group[:split_point]
            cluster_2 = group[split_point:]

            # Assign new cluster IDs
            for member in cluster_1:
                member['clusterId'] = f"{cluster_id}a"
            for member in cluster_2:
                member['clusterId'] = f"{cluster_id}b"

            # Add modified members to the new list
            new_members.extend(cluster_1)
            new_members.extend(cluster_2)

        else:
            # If the cluster has 6 or fewer members, add them as-is
            new_members.extend(group)

    return new_members

def get_smallest_team_index(teams):
    # Calculate the size of each team
    team_sizes = [len(team) for team in teams]
    # Find the smallest size
    min_size = min(team_sizes)
    # Get indices of teams with the smallest size
    smallest_indices = [i for i, size in enumerate(team_sizes) if size == min_size]
    # Randomly choose one of the smallest teams
    return random.choice(smallest_indices)

def distribute():
    global teams, people
    # Repeat until a valid distribution is found
    while True:
        # Step 1: Create n empty team lists (n = number_of_teams)
        teams = [[] for _ in range(number_of_teams)]

        # Step 2: Split large clusters if needed
        new_people = split_large_clusters(people)
        people = new_people

        # Step 3: Group people by clusterId
        clusters = defaultdict(list)
        for person in people:
            clusters[person['clusterId']].append(person)

        # Step 4: Shuffle clusters for randomness
        cluster_list = list(clusters.values())
        random.shuffle(cluster_list)

        # Step 5: Assign clusters to the smallest team
        for cluster in cluster_list:
            # Find the smallest team index
            smallest_team_index = get_smallest_team_index(teams)
            # Assign the cluster to the smallest team
            teams[smallest_team_index].extend(cluster)

        # Step 6: Check if the distribution is valid
        if feasibility_check(teams):
            break  # Exit the loop when a valid distribution is found


def feasibility_check(teams):
    # Check team size constraints
    team_sizes = [len(team) for team in teams]
    max_team_size = max(team_sizes)
    min_team_size = min(team_sizes)

    # Condition A: If the size difference between largest and smallest teams exceeds 7
    if max_team_size - min_team_size > 6:
        return False

    for team in teams:
        # Check if the team is empty
        if not team:
            return False

        # Condition B: Check for at least 3 different dashIds
        unique_dash_ids = {member["dashId"] for member in team}
        if len(unique_dash_ids) < 3:
            return False

    return True

def calculate_team_size_penalty(teams):
    team_sizes = [len(team) for team in teams]
    s_max = max(team_sizes)  # maximum team size
    N = len(teams)

    # Calculate the smoothing index for team size penalty
    size_penalty = math.sqrt(sum((s_i - s_max) ** 2 for s_i in team_sizes) / N)
    return size_penalty

def calculate_gender_ratio_penalty(teams):
    ratio_penalty = 0
    for team in teams:
        # Count males and females
        gender_counts = Counter(person['gender'] for person in team)
        males = gender_counts.get('male', 0)
        females = gender_counts.get('female', 0)

        # Calculate ratio and penalty
        total = males + females
        if total > 0:
            ratio = abs((males / total) - 0.5) * 2  # deviation from 50/50
            ratio_penalty += ratio * total  # weight penalty by team size
    return ratio_penalty

def calculate_dash_id_penalty(teams):
    unique_dash_counts = [len(set(person['dashId'] for person in team)) for team in teams]
    max_unique_dash_ids = max(unique_dash_counts)  # maximum unique dash IDs in a team
    N = len(teams)

    # Calculate smoothing index for dash IDs
    dash_id_penalty = math.sqrt(sum((unique_dash_count - max_unique_dash_ids) ** 2 for unique_dash_count in unique_dash_counts) / N)
    return dash_id_penalty

def fitness(teams):
    # Calculate each penalty factor
    team_size_penalty = calculate_team_size_penalty(teams)
    gender_ratio_penalty = calculate_gender_ratio_penalty(teams)
    dash_id_penalty = calculate_dash_id_penalty(teams)

    # Combine penalties for final fitness score
    fitness_score = team_size_penalty + gender_ratio_penalty + dash_id_penalty
    return fitness_score

def generate_solutions():
    global teams, best_teams_so_far, best_fitness_so_far, max_iterations, final_temp, initial_temp

    i = 0
    t = initial_temp

    distribute()  # Distribute people into teams initially

    best_teams_so_far = deepcopy(teams)  # Use deepcopy to ensure no reference issues
    best_fitness_so_far = fitness(teams)  # Calculate fitness of initial distribution

    current_teams = deepcopy(best_teams_so_far)  # Track current best teams
    current_fitness = best_fitness_so_far  # Track current best fitness

    while i <= max_iterations and t >= final_temp:
        teams = []  # Reset teams

        distribute()  # Distribute people into new teams
        new_fitness = fitness(teams)  # Calculate fitness of new distribution

        delta_fitness = new_fitness - current_fitness

        if delta_fitness < 0:
            current_fitness = new_fitness
            current_teams = deepcopy(teams)
        else:
            r = np.random.rand()
            p = np.exp(-delta_fitness / t)
            if r < p:
                current_fitness = new_fitness
                current_teams = deepcopy(teams)

        # Update best solution if current fitness is better
        if current_fitness < best_fitness_so_far:
            best_teams_so_far = deepcopy(current_teams)
            best_fitness_so_far = current_fitness

        # Cooling schedule (linear or exponential)
        if cooling_type == "linear":
            beta = (initial_temp - final_temp) / max_iterations
            t = initial_temp - beta * i
        else:
            alpha = np.random.uniform(0.7, 0.95)
            t = initial_temp * (alpha ** i)

        i += 1



def export_teams_to_excel():
    global teams
    max_team_size = max(len(team) for team in teams)

    data = {}
    cluster_ids = set()

    for i in range(len(teams)):
        team_data = {
            'Name': [],
            'Dash ID': [],
            'Cluster ID': [],
            'Gender': [],
            'Type': [],
            'Number': []
        } 


        for person in teams[i]:
            team_data['Name'].append(person['name'])
            team_data['Dash ID'].append(person['dashId'])
            team_data['Cluster ID'].append(person['clusterId'])
            team_data['Gender'].append(person['gender'])
            team_data['Type'].append(person['type'])
            team_data['Number'].append(str(person['number']))
            cluster_ids.add(person['clusterId'])

        for key in team_data:
            team_data[key] += [''] * (max_team_size - len(team_data[key]))

        for key in team_data:
            data[f'{key} (Team {i + 1})'] = team_data[key]

    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"

    # Set default column width to 15
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = max(adjusted_width, 15)  # Set the width to at least 15

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Assign colors by base cluster ID
    base_clusters = {str(cluster_id).rstrip('abcdefghijklmnopqrstuvwxyz') for cluster_id in cluster_ids}
    cluster_color_map = {base_cluster: PatternFill(start_color="".join([random.choice("0123456789ABCDEF") for _ in range(6)]), end_color="FFFFFF", fill_type="solid") for base_cluster in base_clusters}

    # Apply colors to clusterId and type cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            column_letter = get_column_letter(cell.column)  # Get column letter to identify "Cluster ID" or "Type" columns

            # Apply color to Cluster ID cells
            if "Cluster ID" in ws.cell(row=1, column=cell.column).value:
                base_cluster = str(cell.value).rstrip('abcdefghijklmnopqrstuvwxyz')
                if base_cluster in cluster_color_map:
                    cell.fill = cluster_color_map[base_cluster]

            # Apply color to Type cells for "F" and "S"
            elif "Type" in ws.cell(row=1, column=cell.column).value:
                if cell.value == 'F':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue for F
                elif cell.value == 'S':
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for S

    # Ask the user where to save the file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb.save(file_path)
        os.startfile(file_path)
        
# ----------------web bridge--------------------


class WebBridge(QObject):
    def __init__(self, people):
        super().__init__()
        self.people = people

    @pyqtSlot("QVariant")
    def edge_added(self, edge):

        from_name = edge['from']
        to_name = edge['to']

        if from_name == to_name:
            return  # Prevent self-connections

        from_person = next((p for p in self.people if p['name'] == from_name), None)
        to_person = next((p for p in self.people if p['name'] == to_name), None)

        if from_person and to_person:
            friends = from_person.get('friends', [])

            # Skip if already connected
            if to_name in friends:
                return

            # Replace first empty slot if it exists
            try:
                empty_index = friends.index("")
                friends[empty_index] = to_name
            except ValueError:
                # No empty slot, append normally unidirectionally
                friends.append(to_name)

            from_person['friends'] = friends


    @pyqtSlot("QVariant")
    def edge_removed(self, edge):
        from_name = edge[0]['from']
        to_name = edge[0]['to']

        for person in self.people:
            friends = person.get('friends', [])

            # Remove the connection while keeping the list length <= 2
            # Removes the conncection bidirectionally
            if person['name'] == from_name:
                if to_name in friends:
                    if len(friends) <= 2:
                        idx = friends.index(to_name)
                        friends[idx] = ''
                    else:
                        friends.remove(to_name)

            if person['name'] == to_name:
                if from_name in friends:
                    if len(friends) <= 2:
                        idx = friends.index(from_name)
                        friends[idx] = ''
                    else:
                        friends.remove(from_name)


    @pyqtSlot(str)
    def edge_edited(self, edge_data_json):
        edge_data = json.loads(edge_data_json)
        old_edge = {"from": edge_data["oldEdge"]["from"], "to": edge_data["oldEdge"]["to"]}
        new_edge = {"from": edge_data["newEdge"]["from"], "to": edge_data["newEdge"]["to"]}

        # Remove the old connection
        self.edge_removed([old_edge])

        # Avoid self-connection
        if new_edge["from"] == new_edge["to"]:
            return

        # Add the new connection
        self.edge_added(new_edge)


# ----------------thread--------------------

class TeamsWorkerSignals(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(str)
    error = pyqtSignal(str)

class TeamsWorker(QRunnable):
    def __init__(self, number_of_teams, people):
        super().__init__()
        self.number_of_teams = number_of_teams
        self.people = people
        self.signals = TeamsWorkerSignals()

    def run(self):
        try:
            
            global people  # Make people global
            people = self.people  # Set the global people to our reference
            assign_clusters()
            
            global number_of_teams
            number_of_teams = self.number_of_teams
            generate_solutions()
            
            export_teams_to_excel()
            

            self.signals.finished.emit()
        except Exception as e:
            self.signals.error.emit(str(e))



class GraphWorkerSignals(QObject):
    finished = pyqtSignal(str)

class GraphWorker(QRunnable):
    def __init__(self, people):
        super().__init__()
        self.people = people
        self.signals = GraphWorkerSignals()

    def run(self):

        try:

            assign_clusters()

            net = Network(notebook=False)

            # Add all people as nodes
            for person in self.people:
                name = person["name"]
                net.add_node(name, label=get_first_and_second_names(name))

            # Add edges based on friends list
            for person in self.people:
                name = person["name"]
                for friend in person.get("friends", []):
                    if friend and friend != "None":
                        net.add_edge(name, friend)

            net.toggle_physics(True)
            temp_html = os.path.join(tempfile.gettempdir(), f"graph_{os.getpid()}.html")
            net.save_graph(temp_html)

            with open(temp_html, "r", encoding="utf-8") as f:
                html = f.read()

            custom_js = """
            <style>
                #searchContainer {
                    position: fixed;
                    top: 10px;
                    left: 50%;
                    transform: translateX(-50%);
                    z-index: 1000;
                    background: white;
                    padding: 10px;
                    border-radius: 5px;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.2);
                }
                #searchBox {
                    padding: 5px;
                    margin-right: 5px;
                    border: 1px solid #ccc;
                    border-radius: 3px;
                }
                #searchButton {
                    padding: 5px 10px;
                    background: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    cursor: pointer;
                }
                #searchButton:hover {
                    background: #45a049;
                }
                #mynetwork {
                    width: 100vw !important;
                    height: 100vh !important;
                    position: fixed !important;
                    top: 0 !important;
                    left: 0 !important;
                }
            </style>
            <div id="searchContainer">
                <input type="text" id="searchBox" placeholder="Search for a name">
                <button id="searchButton" onclick="searchNode()">Search</button>
            </div>
            <script type="text/javascript" src="qrc:///qtwebchannel/qwebchannel.js"></script>
            <script type="text/javascript">
                var bridge = null;

                new QWebChannel(qt.webChannelTransport, function(channel) {
                    bridge = channel.objects.bridge;

                    network.on("addEdge", function(params) {
                        var edgeData = {
                            from: params.from,
                            to: params.to
                        };
                        bridge.edge_added(edgeData);
                    });
                });

                network.setOptions({
                    manipulation: {
                        enabled: true,
                        addNode: false,
                        deleteNode: false,
                        addEdge: function (data, callback) {
                            if (bridge) {
                                var edgeData = {
                                    from: data.from,
                                    to: data.to
                                };
                                bridge.edge_added(edgeData);
                            }
                            callback(data);
                        },
                        deleteEdge: function (data, callback) {
                            if (bridge) {
                                var deletedEdges = data.edges.map(function(edgeId) {
                                    var edge = network.body.data.edges.get(edgeId);
                                    return {
                                        from: edge.from,
                                        to: edge.to
                                    };
                                });
                                bridge.edge_removed(deletedEdges);
                            }
                            callback(data);
                        },
                        editEdge: function (data, callback) {
                            if (bridge) {
                                var edge = network.body.data.edges.get(data.id);
                                const oldData = {
                                    from: edge.from,
                                    to: edge.to
                                };

                                const newData = {
                                    from: data.from,
                                    to: data.to
                                };

                                const payload = JSON.stringify({
                                    oldEdge: oldData,
                                    newEdge: newData
                                });

                                bridge.edge_edited(payload);
                            }
                            callback(data);
                        }
                    }
                });

                function searchNode() {
                    var query = document.getElementById("searchBox").value.trim();
                    if (!query) return;

                    var allNodes = network.body.data.nodes.get();
                    var matchingNodes = allNodes
                        .filter(node => node.label.toLowerCase().includes(query.toLowerCase()))
                        .map(node => node.id);

                    if (matchingNodes.length > 0) {
                        network.selectNodes(matchingNodes);
                    } else {
                        alert("No matching nodes found.");
                    }
                }
            </script>
            """


            # Insert custom JavaScript before </body>
            html = html.replace("</body>", custom_js + "\n</body>")

            with open(temp_html, "w", encoding="utf-8") as f:
                f.write(html)

            self.signals.finished.emit(temp_html)

        except Exception as e:
            self.signals.finished.emit("")



# ----------------frontend--------------------

class GraphTransitionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Heads up!")
        self.setFixedSize(400, 400)
        self.setWindowIcon(QIcon("images/cluster.png"))
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)  # Remove question mark
        
        layout = QVBoxLayout()
        
        # Create label for top text
        top_text = QLabel("Better safe than sorry!")
        top_text.setAlignment(Qt.AlignCenter)
        top_text.setStyleSheet("font-size: 14px; font-weight: bold; margin: 10px;")
        
        # Create label for icon/image
        icon_label = QLabel()
        icon_label.setText("🙏") 
        icon_label.setStyleSheet("font-size: 72px;")
        icon_label.setAlignment(Qt.AlignCenter)
        
        # Create label for bottom text
        bottom_text = QLabel("You are highly advised to save your current matches before entering the graph view.\n\n If anything goes wrong, you can always reload the matches file and continue on from there.")
        bottom_text.setAlignment(Qt.AlignCenter)
        bottom_text.setStyleSheet("font-size: 12px; margin: 10px;")
        bottom_text.setWordWrap(True)
        
        # Create button layout
        button_layout = QHBoxLayout()
        
        # Create OK and Cancel buttons
        cancel_button = QPushButton("Cancel")
        ok_button = QPushButton("Enter Graph View")
        
        # Connect buttons to dialog slots
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        
        # Add buttons to button layout in reverse order
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(ok_button)
        
        # Add all widgets to main layout
        layout.addWidget(top_text)
        layout.addWidget(icon_label)
        layout.addWidget(bottom_text)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)

class ImageDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Welcome")
        self.setFixedSize(800, 600)
        self.setWindowIcon(QIcon("images/cluster.png"))
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)  # Remove question mark
        
        layout = QVBoxLayout()
        
        # Create label for top text
        top_text = QLabel("Please select your Excel file to begin.\n\n The excel file should look like this:")
        top_text.setAlignment(Qt.AlignCenter)
        top_text.setStyleSheet("font-size: 14px; font-weight: bold; margin: 10px;")
        
        # Create label for image
        image_label = QLabel()
        pixmap = QPixmap("images/sampleInput2.png")
        pixmap = pixmap.scaled(800, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignCenter)
        
        # Create label for bottom text
        bottom_text = QLabel("The file should contain the following columns:\n\nName, Dash ID, Gender, Type, Number, Friend 1, Friend 2")
        bottom_text.setAlignment(Qt.AlignCenter)
        bottom_text.setStyleSheet("font-size: 12px; margin: 10px;")
        bottom_text.setWordWrap(True)
        
        # Create button layout
        button_layout = QHBoxLayout()
        
        # Create OK and Cancel buttons
        cancel_button = QPushButton("Cancel")
        ok_button = QPushButton("OK")
        
        # Connect buttons to dialog slots
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        
        # Add buttons to button layout in reverse order
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(ok_button)
        
        # Add all widgets to main layout
        layout.addWidget(top_text)
        layout.addWidget(image_label)
        layout.addWidget(bottom_text)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cluster and Form")
        self.setGeometry(200, 200, 1000, 600)
        self.setWindowIcon(QIcon("images/cluster.png")) 
        self.showMaximized()

        # Add message_label as a class variable
        self.message_label = None

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        # Main layout
        self.main_layout = QVBoxLayout(self.central_widget)

        # Search Bar
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText(" 🔍 Search by Name...")
        self.search_bar.textChanged.connect(self.filter_tree)
        self.main_layout.addWidget(self.search_bar)

        # Central area (stacked layout)
        self.central_area = QWidget()
        self.central_layout = QVBoxLayout(self.central_area)
        self.main_layout.addWidget(self.central_area)

        # Table (QTreeWidget)
        self.tree = QTreeWidget()
        self.tree.setColumnCount(5)
        self.tree.setHeaderLabels(["Name", "Friend 1", "Matches 1", "Matches 2", "Friend 2"])
        self.tree.header().setSectionResizeMode(QHeaderView.Stretch)
        self.tree.itemClicked.connect(self.edit_friend_column)
        self.central_layout.addWidget(self.tree)

        # Graph view
        self.graph_view = QWebEngineView()
        self.central_layout.addWidget(self.graph_view)
        self.graph_view.hide()

        # Bottom buttons layout
        self.button_layout = QHBoxLayout()
        self.main_layout.addLayout(self.button_layout)

        # Initial buttons
        self.load_button = QPushButton("📂 Load Excel")
        self.save_button = QPushButton("💾 Save Matches")
        self.graph_button = QPushButton("🌐 Go to Graph")

        self.button_layout.addWidget(self.load_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.graph_button)

        # Graph mode buttons
        self.back_button = QPushButton("🔙 Back to Matches")
        self.teams_button = QPushButton("🧑‍🤝‍🧑 Create Teams")
        self.back_button.hide()
        self.teams_button.hide()
        self.button_layout.addWidget(self.back_button)
        self.button_layout.addWidget(self.teams_button)

        # Connect buttons
        self.load_button.clicked.connect(self.load_excel_file)
        self.save_button.clicked.connect(self.save_matches)
        self.graph_button.clicked.connect(self.show_graph_view)
        self.back_button.clicked.connect(self.show_table_view)
        self.teams_button.clicked.connect(self.create_teams)

        # Disable buttons initially
        self.save_button.setEnabled(False)
        self.graph_button.setEnabled(False)
        self.teams_button.setEnabled(False)

        # Placeholder for people data
        self.people = []
        self.names = []


    def load_excel_file(self):
        # Show image dialog
        dialog = ImageDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
            if file_path:
                df = pd.read_excel(file_path)
                self.people.clear()

                for _, row in df.iterrows():
                    friends = [
                        row['Friend 1'] if not pd.isna(row['Friend 1']) else "",
                        row['Friend 2'] if not pd.isna(row['Friend 2']) else ""
                    ]

                    person = {
                        "name": row['Name'],
                        "dashId": row['Dash ID'],
                        "gender": row['Gender'],
                        "type": row['Type'],
                        "number": row['Number'],
                        "friends": friends,
                        "original_friends": friends.copy(),
                        "clusterId": 1  # Initialize clusterId with a default value
                    }

                    self.people.append(person)

                self.names = extract_names(self.people)
                self.update_tree()

                # Enable buttons
                self.save_button.setEnabled(True)
                self.graph_button.setEnabled(True)
                self.teams_button.setEnabled(True)


    def update_tree(self):
        self.tree.clear()
        for person in self.people:

            friend1 = person["original_friends"][0]
            friend2 = person["original_friends"][1]
            # Find matches for each friend
            friend1_matches = get_matches(person["friends"][0], self.names)
            friend2_matches = get_matches(person["friends"][1], self.names)

            

            item = QTreeWidgetItem([
                person["name"],
                friend1,
                friend1_matches[0] if friend1_matches else "", 
                friend2_matches[0] if friend2_matches else "", 
                friend2
            ])

            self.tree.addTopLevelItem(item)

            update_person(person, friend1_matches, friend2_matches)

    def filter_tree(self, text):
        text = text.lower()
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)

            # Reset background color
            for col in range(item.columnCount()):
                item.setBackground(col, QBrush())

            if text:  # Only search and highlight if text is not empty
                match_found = False
                
                # Check each column for a match
                for col in range(item.columnCount()):
                    if text in item.text(col).lower():
                        match_found = True
                        break

                # If match found, highlight the row
                if match_found:
                    for col in range(item.columnCount()):
                        item.setBackground(col, QBrush(QColor("yellow")))

    def edit_friend_column(self, item, column):

        if column not in [2, 3]:  # Only edit "Friend 1" or "Friend 2"
            return

        name = item.text(0)
        current_value = item.text(column)

        # Generate matches
        if current_value.lower() == "none" or current_value.strip() == "":
            matches = ["None"] + sorted(self.names)
        else:
            current_matches = get_matches(current_value, self.names)
            other_names = sorted(list(set(self.names) - set(current_matches)))
            matches = current_matches + other_names

        # Setup dropdown
        combo = QComboBox()
        combo.addItems(matches)
        combo.setCurrentText(current_value if current_value else matches[0])

        self.tree.setItemWidget(item, column, combo)
        combo.setFocus()

        def save_selection():
            selected = combo.currentText()
            item.setText(column, selected)
            self.tree.removeItemWidget(item, column)

            # Update people list
            for p in self.people:
                if p["name"] == name:
                    if column == 2:
                        p["friends"][0] = "" if selected == "None" else selected
                    elif column == 3:
                        p["friends"][1] = "" if selected == "None" else selected

        combo.activated.connect(save_selection)
        


    def save_matches(self):
        try:
            export_data = []
            for person in self.people:
                export_data.append({
                    "Name": person["name"],
                    "Dash ID": person["dashId"],
                    "Gender": person["gender"],
                    "Type": person["type"],
                    "Number": str(person["number"]),
                    "Friend 1": person["friends"][0],
                    "Friend 2": person["friends"][1]
                })

            df = pd.DataFrame(export_data)
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Current Matches", "", "Excel Files (*.xlsx)")
            if not file_path:
                return

            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"

            df.to_excel(file_path, index=False)

            try:
                if sys.platform.startswith('darwin'):
                    os.system(f'open "{file_path}"')
                elif os.name == 'nt':
                    os.startfile(file_path)
            except Exception as open_err:
                print(f"Saved, but couldn't open file automatically: {open_err}")

        except Exception as e:
            print(f"Error: Failed to save matches. Details: {e}")

    def show_graph_view(self):
        # Show custom dialog
        dialog = GraphTransitionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.tree.hide()
            self.graph_view.show()
            self.load_button.hide()
            self.save_button.hide()
            self.graph_button.hide()
            self.back_button.show()
            self.teams_button.show()
            self.search_bar.hide()
            
            self.create_graph()


    def create_graph(self):
        worker = GraphWorker(self.people)
        worker.signals.finished.connect(self.show_graph)
        QThreadPool.globalInstance().start(worker)

    def show_graph(self, temp_html):
        if not temp_html:
            return
        
        # Clean up previous graph view if it exists
        if hasattr(self, 'graph_view') and self.graph_view:
            self.central_layout.removeWidget(self.graph_view)
            self.graph_view.deleteLater()
            self.graph_view = None

        # Create new graph view
        self.graph_view = QWebEngineView()
        self.bridge = WebBridge(self.people)

        self.channel = QWebChannel()
        self.channel.registerObject("bridge", self.bridge)
        self.graph_view.page().setWebChannel(self.channel)

        # Add to layout first
        self.central_layout.addWidget(self.graph_view)

        # Add a text message above the graph
        self.message_label = QLabel("HINT: Try to make the clusters as small as possible.")
        self.message_label.setAlignment(Qt.AlignCenter)
        self.message_label.setFixedHeight(30)  # Set a fixed height for the label
        self.central_layout.insertWidget(0, self.message_label)  # Insert at the top of the layout

        # Load the HTML with a small delay to ensure proper initialization
        QTimer.singleShot(100, lambda: self.graph_view.load(QUrl.fromLocalFile(temp_html)))
        
        self.graph_visible = True
        self.graph_button.setText("🌐 Go to Graph")
        self.graph_button.setEnabled(True)



    def show_table_view(self):
        # Remove the message label if it exists
        if self.message_label:
            self.central_layout.removeWidget(self.message_label)
            self.message_label.deleteLater()
            self.message_label = None

        self.graph_view.hide()
        self.update_tree()
        self.tree.show()
        self.load_button.show()
        self.save_button.show()
        self.graph_button.show()
        self.back_button.hide()
        self.teams_button.hide()
        self.search_bar.show()

    def create_teams(self):
        # Disable the button while processing
        self.teams_button.setEnabled(False)
        self.teams_button.setText("⏳ Creating Teams...")

        # Get number of teams
        number_of_teams = simpledialog.askinteger("Input", "Enter the number of teams required:", minvalue=1)
        if number_of_teams is None:
            self.teams_button.setEnabled(True)
            self.teams_button.setText("🧑‍🤝‍🧑 Create Teams")
            return

        # Create and start the worker
        worker = TeamsWorker(number_of_teams, self.people)
        worker.signals.finished.connect(self.team_creation_finished)
        worker.signals.error.connect(self.team_creation_error)
        QThreadPool.globalInstance().start(worker)

    def team_creation_finished(self):
        self.teams_button.setEnabled(True)
        self.teams_button.setText("🧑‍🤝‍🧑 Create Teams")

    def team_creation_error(self, error_message):
        self.teams_button.setEnabled(True)
        self.teams_button.setText("🧑‍🤝‍🧑 Create Teams")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
